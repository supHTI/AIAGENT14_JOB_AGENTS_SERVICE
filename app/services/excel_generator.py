"""
Excel Generation Service for Cooling Period Reports

This module handles generating Excel files for HR-level and Manager-level reports.
"""

import logging
from io import BytesIO
from collections import defaultdict
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("app_logger")


class ExcelGenerator:
    """Service for generating Excel reports for cooling period tracking"""

    # Define consistent styling
    HEADER_FILL = PatternFill(start_color="438EFC", end_color="438EFC", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
    SUBHEADER_FILL = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    SUBHEADER_FONT = Font(bold=True, color="438EFC", size=11)
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    @staticmethod
    def generate_hr_level_excel(job_title_or_mapping, candidates: List[Dict[str, Any]] = None) -> bytes:
        """
        Generate HR-level Excel report. Supports two usages:
        1) Pass a mapping { job_title: [candidates] } to create one sheet per job title.
        2) Pass job_title (str) and candidates (list) to create a single-sheet report.

        Columns: Candidate Name | Candidate ID | Joining Date | Clawback End Date | Days Remaining
        """
        wb = Workbook()

        # If a mapping is provided, create one sheet per job title
        if isinstance(job_title_or_mapping, dict):
            first = True
            for job_title, job_candidates in job_title_or_mapping.items():
                if first:
                    ws = wb.active
                    first = False
                else:
                    ws = wb.create_sheet()

                sheet_name = (job_title or "Job Details")[:31]
                ws.title = sheet_name

                headers = ["Candidate Name", "Candidate ID", "Joining Date", "Clawback End Date", "Days Remaining"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.fill = ExcelGenerator.HEADER_FILL
                    cell.font = ExcelGenerator.HEADER_FONT
                    cell.alignment = ExcelGenerator.CENTER_ALIGNMENT
                    cell.border = ExcelGenerator.BORDER

                for row_num, candidate in enumerate(job_candidates, 2):
                    ws.cell(row=row_num, column=1).value = candidate.get("candidate_name", "N/A")
                    ws.cell(row=row_num, column=2).value = candidate.get("candidate_id", "N/A")
                    ws.cell(row=row_num, column=3).value = str(candidate.get("joining_date", "N/A"))
                    ws.cell(row=row_num, column=4).value = str(candidate.get("clawback_end_date", "N/A"))
                    ws.cell(row=row_num, column=5).value = candidate.get("days_remaining", 0)

                    for col_num in range(1, 6):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.border = ExcelGenerator.BORDER
                        if col_num == 5:
                            cell.alignment = ExcelGenerator.CENTER_ALIGNMENT
                        else:
                            cell.alignment = ExcelGenerator.LEFT_ALIGNMENT

                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 18
                ws.column_dimensions['E'].width = 15

                # summary
                summary_row = len(job_candidates) + 3
                ws.cell(row=summary_row, column=1).value = "Total Candidates:"
                ws.cell(row=summary_row, column=2).value = len(job_candidates)
                ws.cell(row=summary_row, column=1).font = ExcelGenerator.SUBHEADER_FONT

        else:
            # Single sheet behavior (backwards compatible)
            job_title = str(job_title_or_mapping) if job_title_or_mapping else "Job Details"
            ws = wb.active
            ws.title = job_title[:31]

            headers = ["Candidate Name", "Candidate ID", "Joining Date", "Clawback End Date", "Days Remaining"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = ExcelGenerator.HEADER_FILL
                cell.font = ExcelGenerator.HEADER_FONT
                cell.alignment = ExcelGenerator.CENTER_ALIGNMENT
                cell.border = ExcelGenerator.BORDER

            candidates = candidates or []
            for row_num, candidate in enumerate(candidates, 2):
                ws.cell(row=row_num, column=1).value = candidate.get("candidate_name", "N/A")
                ws.cell(row=row_num, column=2).value = candidate.get("candidate_id", "N/A")
                ws.cell(row=row_num, column=3).value = str(candidate.get("joining_date", "N/A"))
                ws.cell(row=row_num, column=4).value = str(candidate.get("clawback_end_date", "N/A"))
                ws.cell(row=row_num, column=5).value = candidate.get("days_remaining", 0)

                for col_num in range(1, 6):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = ExcelGenerator.BORDER
                    if col_num == 5:
                        cell.alignment = ExcelGenerator.CENTER_ALIGNMENT
                    else:
                        cell.alignment = ExcelGenerator.LEFT_ALIGNMENT

            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 15

            summary_row = len(candidates) + 3
            ws.cell(row=summary_row, column=1).value = "Total Candidates:"
            ws.cell(row=summary_row, column=2).value = len(candidates)
            ws.cell(row=summary_row, column=1).font = ExcelGenerator.SUBHEADER_FONT

        # Convert to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_manager_level_excel(job_title_or_mapping: str, hr_candidates_data: List[Dict[str, Any]] | dict = None) -> bytes:
        """
        Generate Manager-level Excel report with sheets per job title.

        Accepts either:
        - A dict mapping { job_title: [rows] } where each row contains hr_name, hr_email and candidate fields
        - A legacy list `hr_candidates_data` where each item is { hr_name, hr_email, candidates: [...] }

        Each sheet columns: HR Name | HR Email | Candidate Name | Candidate ID | Joining Date | Clawback End Date | Days Remaining
        """
        wb = Workbook()

        # Normalize to a job_title -> rows mapping
        job_mapping = {}
        if isinstance(hr_candidates_data, dict):
            job_mapping = hr_candidates_data
        else:
            # Legacy list: iterate HR buckets and distribute by job_title
            job_mapping = {}
            for hr in hr_candidates_data or []:
                hr_name = hr.get("hr_name", "Unassigned")
                hr_email = hr.get("hr_email", "N/A")
                for candidate in hr.get("candidates", []):
                    job_title = candidate.get("job_title", "Unknown Job")
                    row = {
                        "hr_name": hr_name,
                        "hr_email": hr_email,
                        "candidate_name": candidate.get("candidate_name", "N/A"),
                        "candidate_id": candidate.get("candidate_id", "N/A"),
                        "joining_date": str(candidate.get("joining_date", "N/A")),
                        "clawback_end_date": str(candidate.get("clawback_end_date", "N/A")),
                        "days_remaining": candidate.get("days_remaining", 0),
                    }
                    job_mapping.setdefault(job_title, []).append(row)

        # For each job_title, create one sheet per HR (hr_name). This results in
        # number of sheets == number of HRs across all jobs (grouped by job).
        first = True
        used_sheet_names = set()

        for job_title, rows in job_mapping.items():
            # Group rows by HR name
            hr_groups = defaultdict(list)
            for r in rows:
                hr_name = r.get("hr_name") or "Unassigned"
                hr_groups[hr_name].append(r)

            for hr_name, hr_rows in hr_groups.items():
                # Build a sheet name combining job title and HR name, safely truncated
                base_sheet_name = f"{job_title} - {hr_name}" if job_title else hr_name
                sheet_name = base_sheet_name[:31]

                # Ensure unique sheet name by appending a counter if necessary
                counter = 1
                unique_name = sheet_name
                while unique_name in used_sheet_names:
                    suffix = f" {counter}"
                    allowed = 31 - len(suffix)
                    unique_name = (base_sheet_name[:allowed] + suffix)[:31]
                    counter += 1
                used_sheet_names.add(unique_name)

                if first:
                    ws = wb.active
                    ws.title = unique_name
                    first = False
                else:
                    ws = wb.create_sheet(title=unique_name)

                headers = ["HR Name", "HR Email", "Candidate Name", "Candidate ID", "Joining Date", "Clawback End Date", "Days Remaining"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = header
                    cell.fill = ExcelGenerator.HEADER_FILL
                    cell.font = ExcelGenerator.HEADER_FONT
                    cell.alignment = ExcelGenerator.CENTER_ALIGNMENT
                    cell.border = ExcelGenerator.BORDER

                row_num = 2
                for r in hr_rows:
                    ws.cell(row=row_num, column=1).value = hr_name
                    ws.cell(row=row_num, column=2).value = r.get("hr_email", "N/A")
                    ws.cell(row=row_num, column=3).value = r.get("candidate_name")
                    ws.cell(row=row_num, column=4).value = r.get("candidate_id")
                    ws.cell(row=row_num, column=5).value = r.get("joining_date")
                    ws.cell(row=row_num, column=6).value = r.get("clawback_end_date")
                    ws.cell(row=row_num, column=7).value = r.get("days_remaining", 0)

                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.border = ExcelGenerator.BORDER
                        if col_num == 7:
                            cell.alignment = ExcelGenerator.CENTER_ALIGNMENT
                        else:
                            cell.alignment = ExcelGenerator.LEFT_ALIGNMENT
                    row_num += 1

                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 25
                ws.column_dimensions['C'].width = 25
                ws.column_dimensions['D'].width = 15
                ws.column_dimensions['E'].width = 15
                ws.column_dimensions['F'].width = 18
                ws.column_dimensions['G'].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def generate_completed_cooling_excel(completed_candidates: List[Dict[str, Any]], include_hr_details: bool = False) -> bytes:
        """
        Generate Excel report for candidates with completed cooling periods.
        
        Structure:
        - For HR: Candidate Name | Candidate ID | Joining Date | Cooling End Date
        - For Admin: Candidate Name | Candidate ID | HR Name | Joining Date | Cooling End Date
        
        Args:
            completed_candidates: List of completed candidate data
            include_hr_details: Whether to include HR details (for admin report)
            
        Returns:
            Excel file content as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Completed Cooling"
        
        if include_hr_details:
            headers = ["Candidate Name", "Candidate ID", "HR Name", "HR Email", "Joining Date", "Cooling End Date", "Cooling Period (Days)"]
        else:
            headers = ["Candidate Name", "Candidate ID", "Joining Date", "Cooling End Date", "Cooling Period (Days)"]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = ExcelGenerator.HEADER_FILL
            cell.font = ExcelGenerator.HEADER_FONT
            cell.alignment = ExcelGenerator.CENTER_ALIGNMENT
            cell.border = ExcelGenerator.BORDER
        
        # Write data
        for row_num, candidate in enumerate(completed_candidates, 2):
            if include_hr_details:
                ws.cell(row=row_num, column=1).value = candidate.get("candidate_name", "N/A")
                ws.cell(row=row_num, column=2).value = candidate.get("candidate_id", "N/A")
                ws.cell(row=row_num, column=3).value = candidate.get("hr_name", "N/A")
                ws.cell(row=row_num, column=4).value = candidate.get("hr_email", "N/A")
                ws.cell(row=row_num, column=5).value = str(candidate.get("joining_date", "N/A"))
                ws.cell(row=row_num, column=6).value = str(candidate.get("cooling_end_date", "N/A"))
                ws.cell(row=row_num, column=7).value = candidate.get("cooling_period_days", 0)
            else:
                ws.cell(row=row_num, column=1).value = candidate.get("candidate_name", "N/A")
                ws.cell(row=row_num, column=2).value = candidate.get("candidate_id", "N/A")
                ws.cell(row=row_num, column=3).value = str(candidate.get("joining_date", "N/A"))
                ws.cell(row=row_num, column=4).value = str(candidate.get("cooling_end_date", "N/A"))
                ws.cell(row=row_num, column=5).value = candidate.get("cooling_period_days", 0)
            
            # Apply styling
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = ExcelGenerator.BORDER
                cell.alignment = ExcelGenerator.LEFT_ALIGNMENT
        
        # Adjust column widths
        col_widths = [25, 15, 20, 25, 15, 18, 18] if include_hr_details else [25, 15, 15, 18, 18]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        # Add summary
        summary_row = len(completed_candidates) + 3
        ws.cell(row=summary_row, column=1).value = "Total Completed:"
        ws.cell(row=summary_row, column=2).value = len(completed_candidates)
        ws.cell(row=summary_row, column=1).font = ExcelGenerator.SUBHEADER_FONT
        
        # Convert to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
